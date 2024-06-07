from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from statFunctions import *

# Définition des chemins d'accès aux images dans un tableau numpy
images = np.array([
    ['Ressources/GR_300_307-323_fluo_000.tif', 'Ressources/GRMD_307-323_fluo_000.tif', 'Ressources/GRMDT_307-323_fluo_000.tif'],
    ['Ressources/GR_327-353_fluo_000.tif', 'Ressources/GRMD_327-353_fluo_000.tif', 'Ressources/GRMDT_327-353_fluo_000.tif'],
    ['Ressources/GR_370-410_fluo_000.tif', 'Ressources/GRMD_370-410_fluo_000.tif', 'Ressources/GRMDT_370-410_fluo_000.tif'],
    ['Ressources/GR_420-480_fluo_000.tif', 'Ressources/GRMD_420-480_fluo_000.tif', 'Ressources/GRMDT_420-480_fluo_000.tif'],
    ['Ressources/GR_435-455_fluo_000.tif', 'Ressources/GRMD_435-455_fluo_000.tif', 'Ressources/GMDT_435-455_fluo_000.tif']
])

# Liste des types de natures pour les images
natures = ['saine', 'malade', 'traite']
# Liste des statistiques à calculer
statList = ["Autocorrélation", "Energie", "Contraste", "Homogénéité", "PIQE", "Corrélation", "Erreur quadratique moyenne", "SSIM", "ODC"]

# Charger le fichier Excel contenant les données statistiques
wb = load_workbook('data.xlsx')


# Fonction pour obtenir les données statistiques pour une image donnée
def getDataForImage(selected_path, selected_stats):
    data = {}
    for image_type, image_path in selected_path.items():
        # Trouver l'index de l'image dans le tableau 'images'
        image_index = np.where(images == image_path)
        row_index = image_index[0][0] + 1
        col_index = get_column_letter(image_index[1][0] + 1)
        for stat, statIndex in selected_stats.items():
            # Accéder à la feuille Excel correspondant à la statistique
            sheet = wb[stat]
            # Obtenir la valeur de la statistique depuis le fichier Excel
            statValue = float(sheet[col_index + str(row_index)].value)
            data[stat] = statValue
    return data

# Fonction pour obtenir les données statistiques pour plusieurs images
def getDataForImages(selected_paths, selected_stats):
    data = {}
    images_indexes = []
    for path in selected_paths.values():
        # Trouver l'index de l'image dans le tableau 'images'
        image_row = np.where(images == path)[0][0]
        image_column = np.where(images == path)[1][0]
        images_indexes.append([image_row, image_column])
    
    row_index = (images_indexes[0][0] * 3 + 1) + images_indexes[0][1]
    column_index = get_column_letter(images_indexes[1][1] + 1)
    for stat, statIndex in selected_stats.items():
        # Accéder à la feuille Excel correspondant à la statistique
        sheet = wb[stat]
        # Obtenir la valeur de la statistique depuis le fichier Excel
        statValue = float(sheet[column_index + str(row_index)].value)
        data[stat] = statValue
    return data
